import datetime
import json
import os 
import random

from openpyxl.workbook import Workbook
from openpyxl.styles import Font

from django import utils
from django.core import serializers
from django.shortcuts import render, redirect, get_object_or_404, HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.forms.models import model_to_dict
from django.core.files.storage import FileSystemStorage

from django.http import JsonResponse

from .models import Event, Person, RelatedPerson, AdressesPlacesOfBirth, AdressesPlacesOfLive, AdressesPlacesOfWork, OtherAdresses, Changes, ChangesEvent, FilesPerson, FilesEvent
from .models import Entity, Division, Filial, Representation
from .models import Logging, LoggingUser
from .forms import PersonForm, EventForm, AdressBirthForm, AdressWorkForm, FilterPersonForm, FilterEventForm, AdressLiveForm, AdressOtherForm
from .forms import EMPTY_VALUE

# Views for view data


# login/logout


def loginView(request):
    warning = None
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(username=username, password=password)
        if user is not None:
            login(request, user)
            LoggingUser(author=user.last_name+' '+user.first_name, type='Вход').save()
            return JsonResponse(data={'status': 200})
        else:
            warning = 'Введен неправильный логин или пароль!'
            return JsonResponse(data={'status': 500, 'warning': warning})
    return render(request, 'login.html')

def logoutView(request):
    LoggingUser(author=request.user.last_name+' '+request.user.first_name, type='Выход').save()
    logout(request)
    return redirect('/login')


# main

@login_required
def startPage(request):
    return render(request, 'main-page.html')

@login_required
def eventsView(request):
    form = FilterEventForm(request.GET)
    events = Event.objects.order_by('-id')

    for i, j in request.GET.items():
        if i in ('filial', 'division', 'type', 'entity'):
            if j:
                events = events.filter(**{i: j})
        if i == 'search_input':
            if j:
                m = j.split()
                events_new = []
                for event in events:
                    if all(x.lower() in event.description.lower() for x in m):
                        events_new.append(event)
                events = events_new

    return render(request, 'home.html', {'events': events,
                                        'name': 'События',
                                        'name_number': 'События',
                                        'event': 1,
                                        'form': form,
                                        'author': request.user.first_name + ' ' + request.user.last_name})

@login_required
def personsView(request):
    form = FilterPersonForm(request.GET)
    print(request.GET)
    persons = Person.objects.order_by('-id')
    if 'role' in request.GET:
        if request.GET['role']:
            persons = list(set(get_object_or_404(Person, id=i.id_person) for i in RelatedPerson.objects.filter(role=request.GET['role'])))
    for i, j in request.GET.items():
        if not j:
            continue
        if i == 'role':
            continue
        elif i in ('sex', 'entity', 'division', 'filial', 'representation'):
            if i == 'sex':
                m = j
                persons = filter(lambda x: x.sex==m, persons)
            else:
                m = j
                filters = {i: m}
                persons = filter(lambda x: AdressesPlacesOfWork.objects.filter(id_place_of_work=x.id, **filters), persons)
        if i == 'search_input':
            m = j.split()
            persons_new = []
            for person in persons:
                if all(x.lower() in person.fio.lower() for x in m):
                    persons_new.append(person)
            persons = persons_new
    
    return render(request, 'home.html', {'person': persons, 
                                        'name_number': 'Лица', 
                                        'name': 'Лица', 
                                        'person_num': 1,
                                        'author': request.user.first_name + ' ' + request.user.last_name,
                                        'form': form})


# Views for add data

@login_required
def addEvent(request):
    header = 'Заполните данные о событии'
    if request.method == 'POST':
        form = EventForm(data=request.POST)
        if form.is_valid():
            user = User.objects.get(username=request.user)
            author = user.last_name + ' ' + user.first_name
            value = form.save(commit=False)
            value.author = author
            value.save()

            Logging(author=author, type='Создание события', logged_event='{} создал запись события {}'.format(author, value.id)).save()

            return redirect('/event/'+str(value.id))
    else:
        form = EventForm()
    return render(request, 'add-event.html', {'form': form,
                                              'header': header,
                                              'events': 1})


@login_required
def addPerson(request, redirected=None):
    header = 'Заполните данные о человеке'
    if request.method == 'POST':
        author = request.user.last_name + ' ' + request.user.first_name
        if 'data' in request.POST:
            try:
                fio, birthday = request.POST['data'].split(' | ')
                person = get_object_or_404(Person, fio=fio, birthday=birthday)
                event_id = request.POST['event_id']
                role = request.POST['role']
                if person:
                    related = RelatedPerson(id_event=event_id, role=role, id_person=person.id)
                    if role == 'Другой фигурант':
                        related.role_desc = request.POST['role_desc']
                        related.save()
                        person.id_related = related.id
                    related.save()
                    person.save()

                    Logging(author=author, type='Создание отношения', logged_event='{} создал запись отношения к событиям {}:\nЛицо с ID {} теперь {}'.format(author, related.id, person.id, role)).save()

                    return JsonResponse({'status': '200'})
            except Exception as e:
                return JsonResponse({'status': '404'})
        else:
            form = PersonForm(data=request.POST)
            if form.is_valid():
                fio_p = form.cleaned_data['last_name'] + ' ' + form.cleaned_data['first_name'] + ' ' + form.cleaned_data['second_name']
                birthday = form.cleaned_data['birthday']

                try:
                    if Person.objects.get(fio=fio_p, birthday=birthday):
                        return JsonResponse(data={'message': 'Такой человек уже существует!'})
                except Exception as e:
                    print(e)

                value = form.save(commit=False)
                value.author = author
                value.fio = fio_p
                value.save()

                Logging(author=author, type='Создание лица', logged_event='{} создал запись лица {}'.format(author, value.id)).save()

                if 'redirected' in request.POST:
                    return JsonResponse({'redirect': '200'})
                else:
                    return JsonResponse({'status': '200', 'id': value.id})      
    else:
        form = PersonForm()
    if redirected:
        return render(request, 'add-person.html', {'form': form,
                                               'header': header,
                                               'redirected': 1})
    else:
        return render(request, 'add-person.html', {'form': form,
                                               'header': header})


@login_required
def set_adaptive(request):
    if 'name' in request.GET:
        gets = request.GET
        if gets['value'] == '':
            return JsonResponse(data={'data': '500'}) 
        if gets['name'] == 'entity':
            a = [i.name for i in Division.objects.filter(id_entity=Entity.objects.get(name=gets['value']))]
            return JsonResponse(data={'data': a})
        elif gets['name'] == 'division':
            a = [i.name for i in Filial.objects.filter(id_division=Division.objects.get(name=gets['value']))]
            return JsonResponse(data={'data': a})
        elif gets['name'] == 'filial':
            a = [i.name for i in Representation.objects.filter(id_filial=Filial.objects.get(name=gets['value']))]
            return JsonResponse(data={'data': a})
        else:
            return JsonResponse(data={'status': '500'})
    else:
        return redirect(request, 'home')


@login_required
def add_person_with_redirect(request):
    return addPerson(request, True)


@login_required
def addAdress(request, person_id):
    if request.method == 'POST':
        role = request.POST['role']
        user = User.objects.get(username=request.user)
        author = user.first_name + ' ' + user.last_name
        person = Person.objects.get(id=person_id)

        if role == '1':
            form = AdressBirthForm(data=request.POST)
        elif role == '2':
            form = AdressWorkForm(data=request.POST)
        elif role == '3':
            form = AdressLiveForm(data=request.POST)
        else:
            form = AdressOtherForm(data=request.POST)
        if form.is_valid():
            value = form.save(commit=False)
            value.who_added = author

            if role == '1':
                value.id_place_of_birth = person
                header = 'Место рождения'
            elif role == '2':
                value.id_place_of_work = person
                header = 'Место работы'
            elif role == '3':
                value.id_place_of_live = person
                header = 'Адрес регистрации'
            elif role == '4':
                value.id_other_places = person
                header = 'Другой адрес'
                
            value.save()
            data = {i[0]: i[1] for i in [k for k in value.__iter__()][2:-1]}

        Logging(author=author, type='Создание адреса', logged_event='{} создал запись {} с ID: {}'.format(author, header, value.id)).save()

        return JsonResponse(data={'status': 200, 'data': data, 'header': header})
    else:
        form_other = AdressOtherForm()
        form_live = AdressLiveForm()
        form_birth = AdressBirthForm()
        form_work = AdressWorkForm()


    return render(request, 'add-adress.html', {
        'form_live': form_live,
        'form_other': form_other,
        'form_birth': form_birth,
        'form_work': form_work,
        'person_id': person_id
    })


@login_required
def add_person_on_event(request, id):
    if request.user.is_superuser or (request.user.last_name+' '+request.user.first_name == Event.objects.get(id=id).author):
        form = PersonForm()
        return render(request, 'add-person-on-event.html', {'form_person': form,
                                                            'event_id': id})
    else:
        return redirect('home')
   



# VIEWS FOR EDIT

@login_required
def edit_person(request, id_person):
    if not request.user.is_superuser and not (request.user.last_name+' '+request.user.first_name == Person.objects.get(id=id_person).author):
        return redirect('home')
    if request.method == 'POST':
        desc = request.POST.get('description')
        person = get_object_or_404(Person, id=id_person)
        change_value = Changes(id_user=request.user.id, edit_from=person.description, edit_to=desc, type_edit='Описание', id_record=person.id, record_name='Лицо')
        change_value.save()

        person.description = desc
        person.add_at = utils.timezone.now()

        person.save()

        author = request.user.last_name+' '+request.user.first_name
        Logging(author=author, type='Изменение лица', logged_event='{} изменил запись лица {}'.format(author, person.id)).save()

        return redirect(f'/person/{id_person}')
    else:
        model = model_to_dict(Person.objects.get(id=id_person))
        model['birthday'] = datetime.datetime.strftime(model['birthday'], '%Y-%m-%d')
        form = PersonForm(model)
    return render(request, 'edit.html', {'form': form,
                                         'header': 'лица',
                                         'person': 1})


@login_required
def edit_event(request, id_event):
    if not request.user.is_superuser and not (request.user.last_name+' '+request.user.first_name == Event.objects.get(id=id_event).author):
        return redirect('home')
    if request.method == 'POST':

        author = request.user.last_name+' '+request.user.first_name

        edit_from = json.dumps(model_to_dict(Event.objects.get(id=id_event)), ensure_ascii=False, default=str)
        event = get_object_or_404(Event, id=id_event)
        description = request.POST['description']
        way = request.POST['way']
        instrument = request.POST['instrument']
        relation = request.POST['relation']
        object = request.POST['object']
        place = request.POST['place']

        event.description = description
        event.way = way
        event.instrument = instrument
        event.relation = relation
        event.object = object
        event.place = place

        event.save()

        edit_to = json.dumps(model_to_dict(Event.objects.get(id=id_event)), ensure_ascii=False, default=str)

        if not edit_from == edit_to:
            change_value = ChangesEvent(id_user=request.user.id, edit_from=edit_from, edit_to=edit_to, id_record=id_event, record_name='Событие')
            change_value.save()

        Logging(author=author, type='Изменение события', logged_event='{} изменил запись события {}'.format(author, event.id)).save()

        return redirect(f'/event/{id_event}')
    else:
        model = model_to_dict(Event.objects.get(id=id_event))
        model['date_incedent'] = datetime.datetime.strftime(model['date_incedent'], '%Y-%m-%d %H:%M')
        form = EventForm(model)
    return render(request, 'edit.html', {'form': form,
                                         'header': 'события',
                                         'event': 1})








# /search/?term=

@login_required
def searchPerson(request):
    person = request.GET.get('term')
    persons = []

    if person:
        intruders_list = Person.objects.filter(fio__iregex=person)
        for obj in intruders_list:
            persons.append(obj.fio + ' | ' + str(obj.birthday))

    return JsonResponse({'status': 200, 'data': persons}, json_dumps_params={'ensure_ascii': False}, content_type='application/json; charset=utf-8')









# Views for view event/person

@login_required
def eventView(request, event_id):
    if request.method == 'POST' and request.FILES:
        event = Event.objects.get(id=event_id)

        if 'file' in request.FILES:
            file = request.FILES['file']
            filename = file.name
            value = FilesEvent(event_id=event_id, file=file, filename=filename)
            value.save()


        event.save()
        return redirect('../event/{}'.format(event_id))
    
    files = FilesEvent.objects.filter(event_id=event_id)

    event = Event.objects.get(id=event_id)
    persons = RelatedPerson.objects.filter(id_event=event_id)
    author = request.user.last_name + ' ' + request.user.first_name

    injureds = [Person.objects.get(id=i.id_person) for i in persons if i.role == 'Потерпевший']
    witnesses = [Person.objects.get(id=i.id_person) for i in persons if i.role == 'Свидетель']
    intruders = [Person.objects.get(id=i.id_person) for i in persons if i.role == 'Нарушитель']
    others = [{'i': Person.objects.get(id=i.id_person), 'j': i.role_desc} for i in persons if i.role == 'Другой фигурант']

    persons_list = [{'persons': intruders, 'name': 'Нарушители'},
                    {'persons': witnesses, 'name': 'Свидетели'},
                    {'persons': injureds, 'name': 'Потерпевшие'},
                    {'persons': others, 'name': 'Остальные фигуранты'}
               ]

    if request.user.username == 'root':
        changes = ChangesEvent.objects.filter(id_record=event_id)
    else:
        changes = ChangesEvent.objects.filter(id_record=event_id, id_user=request.user.id)

    return render(request, 'event.html', {'event': event, 
                                          'persons_list': persons_list,
                                          'event_id': event_id,
                                          'changes': changes,
                                          'files': files,
                                          'author': author})

@login_required
def personView(request, person_id):
    if request.method == 'POST' and request.FILES:
        person = Person.objects.get(id=person_id)

        if 'image' in request.FILES:
            image = request.FILES['image']
            person.person_image = image
        else:
            file = request.FILES['file']
            filename = file.name
            value = FilesPerson(person_id=person_id, file=file, filename=filename)
            value.save()


        person.save()
        return redirect('../person/{}'.format(person_id))

    author = request.user.last_name + ' ' + request.user.first_name

    person = Person.objects.get(id=person_id)

    image = person.person_image
    files = FilesPerson.objects.filter(person_id=person_id)
    events = RelatedPerson.objects.filter(id_person=person_id)

    try:
        places_of_birth = AdressesPlacesOfBirth.objects.filter(id_place_of_birth=person.id)
        places_of_work = AdressesPlacesOfWork.objects.filter(id_place_of_work=person.id)
        places_of_live = AdressesPlacesOfLive.objects.filter(id_place_of_live=person.id)
        other_places = OtherAdresses.objects.filter(id_other_places=person.id)
    except:
        places_of_birth = None
        places_of_work = None
        places_of_live = None
        other_places = None

    places = [{'place': places_of_birth, 'name': 'Место рождения:'}, 
              {'place': places_of_work, 'name': 'Место работы:'},
              {'place': places_of_live, 'name': 'Адрес регистрации:'},
              {'place': other_places, 'name': 'Остальные адреса:'}]

    if request.user.username == 'root':
        changes = Changes.objects.filter(id_record=person_id)
    else:
        changes = Changes.objects.filter(id_record=person_id, id_user=request.user.id)

    table_name = 'Лицо'

    return render(request, 'person_info.html', {'person': person, 'events': events, 'table_name': table_name,
                                                'places': places,
                                                'changes': changes,
                                                'image': image,
                                                'files': files,
                                                'author': author})


@login_required
def changeView(request, change_id):
    if request.user.is_superuser:
        change = get_object_or_404(ChangesEvent, id=change_id)
        change_from = json.loads(change.edit_from).items
        change_to = json.loads(change.edit_to).items
        return render(request, 'change-page.html', {'change_from': change_from,
                                                    'change_to': change_to})
    else:
        return redirect('home')


# REPORT

def create_report(request):
    event_form = FilterEventForm()
    person_form = FilterPersonForm()
    return render(request, 'create-report.html', {
        'event_form': event_form,
        'person_form': person_form,
    })

    # author = request.user.last_name + ' ' + request.user.first_name

    # def counting_by(by:str, num:int, name:str, rule=None):
    #     ws['B'+str(num)] = 'Количество событий по {}'.format(name)
    #     ws['B'+str(num)].font = Font(bold=True)
    #     if rule == 'work':
    #         value = Event.objects.all().values_list(by, flat=True).distinct()
    #     num += 1

    #     for i in value:
    #         ws['B'+str(num)] = i
    #         ws['C'+str(num)] = value.filter(**{by: i}).count()
    #         num += 1
    #     return num

    # start = datetime.datetime.strptime(request.GET['start'], '%Y-%m-%d').date()
    # end_real = datetime.datetime.strptime(request.GET['end'], '%Y-%m-%d').date()
    # end = datetime.datetime.strptime(request.GET['end'], '%Y-%m-%d').date() + datetime.timedelta(days=1)

    # flag = False
    # if 'sex' in request.GET:
    #     sex = request.GET['sex']
    #     role = request.GET['role']
    #     flag = True
    # entity = request.GET['entity']
    # division = request.GET['division']
    # filial = request.GET['filial']

    # if not flag:
    #     events = Event.objects.filter(date_incedent__range=(start, end), entity__icontains=entity, division__icontains=division, filial__icontains=filial)
    # else:
    #     events = Person.objects.filter(sex__icontains=sex, add_at__range=(start, end))
    #     rel = RelatedPerson.objects.filter(role__icontains=role)
    #     ent = AdressesPlacesOfWork.objects.filter(entity__icontains=entity)
    #     ent = ent.filter(locality__icontains=filial)
    #     # ent = ent.filter(filial__icontains=filial)
    #     id_person = [int(i) for i in rel.values_list('id_person', flat=True).distinct()]
    #     id_place = [int(i) for i in ent.values_list('id_place_of_work', flat=True).distinct()]
    #     events = events.filter(id__in=id_person)
    #     events = events.filter(id__in=id_place)

    # if not events:
    #     return JsonResponse(data={'status': '500'})
    
    # start_rus = start.strftime('%d.%m.%Y')
    # end_rus = end_real.strftime('%d.%m.%Y')

    # wb = Workbook()
    # ws = wb.active
    # ws.column_dimensions['B'].width = 50
    # num = 2


    # # Title
    # ws['B'+str(num)] = 'Отчет за период: ' +  start_rus + ' - ' + end_rus


    # # Main count
    # num += 2
    # ws['B'+str(num)] = 'Общее количество событий:'
    # ws['B'+str(num)].font = Font(bold=True)
    # ws['C'+str(num)] = len(events)
    # num += 2

    # # Count by entity
    # if entity == '' and division == '' and filial == '':
    #     num = counting_by('entity', num, 'юридическим лицам', 'work')

    # num += 1

    # # Count by division
    # if division == '' and filial == '':
    #     num = counting_by('division', num, 'дивизионам', 'work')

    # num += 1

    # # Count by filials
    # if filial == '':
    #     num = counting_by('filial', num, 'филиалам', 'work')

    # # Count by type
    

    # if not flag:
    #     type = 'События'
    # else:
    #     type = 'Лица'

    # if not flag:
    #     num += 1
    #     ws['B'+str(num)] = 'Количество событий по виду'
    #     ws['B'+str(num)].font = Font(bold=True)
    #     types = events.values_list('type', flat=True).distinct()
    #     num += 1
    #     for i in types:
    #         ws['B'+str(num)] = i
    #         ws['C'+str(num)] = events.filter(type=i).count()
    #         num += 1

    # if author not in os.listdir(path='reports/'):
    #     os.mkdir(path='reports/' + author)
    #     os.mkdir(path='reports/{}/События'.format(author))
    #     os.mkdir(path='reports/{}/Лица'.format(author))
    
    # wb.save('reports/' + '{}/'.format(request.user.last_name + ' ' + request.user.first_name) + '{}/'.format(type) + 'Отчет по событиям ' + start_rus + ' - ' + end_rus + '.xlsx')


    # return JsonResponse(data={'status': 200})


# HANDLING ERRORS

def handling_404(request, exceptions):
    return render(request, '404.html')


# tests

@login_required
def testJson(request, event_id):
    person_data = {'name': 'Alex', 'date': '10-02-2002'}
    data = request.GET.dict()
    if not data:
        data['persons'] = []
        data['persons'].append(request.GET)
    else:
        data['persons'].append(data)
    
    return JsonResponse(data=data, json_dumps_params={'ensure_ascii': False}, content_type='application/json; charset=utf-8')

def create_random():
    a = ['Хакимов', 'Ананьев', 'Гринцов', 'Рамтышев', 'Ватутин', 'Коломеец']
    b = ['Виктор', 'Александр', 'Степан', 'Владимир', 'Антон', 'Владислав']
    c = ['Константинович', 'Владимирович', 'Алексеевич', 'Степанович', 'Александрович', 'Викторович']
    day = [i for i in range(1, 28)]
    month = [i for i in range(1, 13)]
    year = [i for i in range(1980, 2003)]


    for olt in range(100):
        try:
            f = random.choice(a)
            n = random.choice(b)
            o = random.choice(c)
            value = Person(fio=f+' '+n+' '+o, first_name=n, last_name=f, second_name=o, description='-', sex='м', birthday=datetime.date(random.choice(year), random.choice(month), random.choice(day)), author='Хакимов Виктор')
            value.save()
        except Exception as e:
            print(e)
            continue
